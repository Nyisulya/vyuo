"""
export_combined_excel.py
========================
Script inayounganisha data kutoka vitabu vitatu:
  1. tcu/tcu_data.xlsx        → TCU (Degree/University programs)
  2. nacte/nacte_data.xlsx    → NACTE Book 1 (Certificate/Diploma)
  3. nacte/nacte_data2.xlsx   → NACTE Book 2 (Certificate/Diploma)

Matokeo: combined_data.xlsx (ina sheets 4):
  - TCU        → data ya TCU tu
  - NACTE_1    → data ya NACTE Book 1
  - NACTE_2    → data ya NACTE Book 2
  - ALL_DATA   → data yote pamoja (merged)

Unaweza kuendesha: python export_combined_excel.py
"""

import os
import sys
import pandas as pd
from datetime import datetime
from pathlib import Path

# ─── Mipangilio ya njia za faili ────────────────────────────────────────────
BASE_DIR = Path(__file__).parent

TCU_FILE   = BASE_DIR / "tcu" / "tcu_data.xlsx"
NACTE1_FILE = BASE_DIR / "nacte" / "nacte_data.xlsx"
NACTE2_FILE = BASE_DIR / "nacte" / "nacte_data2.xlsx"
OUTPUT_FILE = BASE_DIR / "combined_data.xlsx"

# ─── Rangi za header kwa kila sheet ─────────────────────────────────────────
HEADER_COLORS = {
    "TCU":      "1F4E79",  # Bluu ya giza
    "NACTE_1":  "375623",  # Kijani ya giza
    "NACTE_2":  "5C3317",  # Kahawia
    "ALL_DATA": "4B0082",  # Indigo
}

def load_tcu(filepath: Path) -> pd.DataFrame:
    """Pakia data ya TCU na ongeza safu 'Source'."""
    if not filepath.exists():
        print(f"  [ONYO] Faili la TCU halipatikani: {filepath}")
        return pd.DataFrame()
    
    df = pd.read_excel(filepath)
    df["Source"] = "TCU"
    df["Category"] = "Degree"
    
    # Standardize column names
    rename_map = {}
    for col in df.columns:
        col_lower = col.lower()
        if "university" in col_lower and "University" not in rename_map.values():
            rename_map[col] = "University"
        elif "programme" in col_lower or "program" in col_lower:
            rename_map[col] = "Programme"
        elif "code" in col_lower:
            rename_map[col] = "Code"
        elif "requirement" in col_lower:
            rename_map[col] = "Requirements"
        elif "duration" in col_lower:
            rename_map[col] = "Duration"
    df = df.rename(columns=rename_map)
    
    # Ensure standard columns exist
    for col in ["University", "Programme", "Code", "Requirements", "Duration", "Source", "Category"]:
        if col not in df.columns:
            df[col] = ""
    
    return df[["University", "Programme", "Code", "Requirements", "Duration", "Source", "Category"]]


def load_nacte(filepath: Path, book_num: int) -> pd.DataFrame:
    """Pakia data ya NACTE na ongeza safu 'Source'."""
    if not filepath.exists():
        print(f"  [ONYO] Faili la NACTE{book_num} halipatikani: {filepath}")
        return pd.DataFrame()
    
    df = pd.read_excel(filepath)
    df["Source"] = f"NACTE_Book{book_num}"
    df["Category"] = "Certificate/Diploma"
    
    # Standardize column names
    rename_map = {}
    for col in df.columns:
        col_lower = col.lower()
        if "university" in col_lower or "institution" in col_lower:
            rename_map[col] = "University"
        elif "programme" in col_lower or "program" in col_lower:
            rename_map[col] = "Programme"
        elif "requirement" in col_lower:
            rename_map[col] = "Requirements"
        elif "duration" in col_lower:
            rename_map[col] = "Duration"
        elif "capacity" in col_lower:
            rename_map[col] = "Capacity"
        elif col_lower == "fee":
            rename_map[col] = "Fee"
        elif "region" in col_lower:
            rename_map[col] = "Region"
        elif "ownership" in col_lower or "umiliki" in col_lower:
            rename_map[col] = "Ownership"
    df = df.rename(columns=rename_map)
    
    # Ensure standard columns
    for col in ["University", "Programme", "Requirements", "Duration", "Capacity", "Fee", "Region", "Ownership", "Source", "Category"]:
        if col not in df.columns:
            df[col] = ""
    
    return df[["University", "Programme", "Requirements", "Duration", "Capacity", "Fee", "Region", "Ownership", "Source", "Category"]]


def style_worksheet(worksheet, header_color: str, df: pd.DataFrame):
    """Weka rangi na mtindo wa header na rows."""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Header style
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align   = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # Weka urefu wa safu (auto-fit)
        for col_idx, col_name in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)
            if len(df) > 0:
                col_max = df[col_name].fillna("").astype(str).str.len().max()
                col_max = int(col_max) if col_max == col_max else 0  # NaN check
            else:
                col_max = 0
            max_len = min(max(len(str(col_name)), col_max), 60)  # Usizidi 60 chars
            worksheet.column_dimensions[col_letter].width = max_len + 4
        
        # Style ya header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        
        worksheet.row_dimensions[1].height = 30
        
        # Weka urefu wa mstari + alternating rows
        alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
            for cell in row:
                cell.alignment = left_align
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
            worksheet.row_dimensions[row_idx].height = 20
        
        # Freeze pane baada ya header
        worksheet.freeze_panes = "A2"
        
        # Auto filter kwenye header
        worksheet.auto_filter.ref = worksheet.dimensions
        
    except ImportError:
        print("  [ONYO] openpyxl haijapatikana kwa styling. Data itahifadhiwa bila mtindo.")


def main():
    print("=" * 60)
    print("  KUUNGANISHA DATA KUTOKA VITABU VITATU")
    print(f"  Tarehe: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    
    # ── Pakia data ──────────────────────────────────────────────
    print("\n[1/4] Kupakia data ya TCU...")
    df_tcu = load_tcu(TCU_FILE)
    print(f"       Rekodi za TCU: {len(df_tcu):,}")
    
    print("[2/4] Kupakia data ya NACTE Book 1...")
    df_nacte1 = load_nacte(NACTE1_FILE, 1)
    print(f"       Rekodi za NACTE_1: {len(df_nacte1):,}")
    
    print("[3/4] Kupakia data ya NACTE Book 2...")
    df_nacte2 = load_nacte(NACTE2_FILE, 2)
    print(f"       Rekodi za NACTE_2: {len(df_nacte2):,}")
    
    # ── Unganisha data yote ──────────────────────────────────────
    print("[4/4] Kuunganisha na kuhifadhi Excel...")
    
    # Merge all into one sheet (ALL_DATA)
    # Align columns - tumia union ya columns zote
    df_all = pd.concat([df_tcu, df_nacte1, df_nacte2], ignore_index=True, sort=False)
    df_all = df_all.fillna("")
    
    # Reorder columns: Source na Category kwanza kwa kuelewa rahisi
    priority_cols = ["Source", "Category", "University", "Programme", "Requirements", "Duration"]
    other_cols = [c for c in df_all.columns if c not in priority_cols]
    df_all = df_all[priority_cols + other_cols]
    
    # Pia reorder kwa TCU
    if len(df_tcu) > 0:
        tcu_cols = ["University", "Programme", "Code", "Requirements", "Duration", "Source", "Category"]
        df_tcu_out = df_tcu.reindex(columns=tcu_cols, fill_value="")
    else:
        df_tcu_out = df_tcu
    
    # ── Andika Excel na sheets nyingi ────────────────────────────
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        
        # Sheet 1: TCU
        if len(df_tcu_out) > 0:
            df_tcu_out.to_excel(writer, sheet_name="TCU", index=False)
            style_worksheet(writer.sheets["TCU"], HEADER_COLORS["TCU"], df_tcu_out)
        else:
            pd.DataFrame(columns=["University","Programme","Code","Requirements","Duration","Source","Category"]).to_excel(
                writer, sheet_name="TCU", index=False)
        
        # Sheet 2: NACTE_1
        if len(df_nacte1) > 0:
            df_nacte1.to_excel(writer, sheet_name="NACTE_1", index=False)
            style_worksheet(writer.sheets["NACTE_1"], HEADER_COLORS["NACTE_1"], df_nacte1)
        else:
            pd.DataFrame().to_excel(writer, sheet_name="NACTE_1", index=False)
        
        # Sheet 3: NACTE_2
        if len(df_nacte2) > 0:
            df_nacte2.to_excel(writer, sheet_name="NACTE_2", index=False)
            style_worksheet(writer.sheets["NACTE_2"], HEADER_COLORS["NACTE_2"], df_nacte2)
        else:
            pd.DataFrame().to_excel(writer, sheet_name="NACTE_2", index=False)
        
        # Sheet 4: ALL_DATA (yote pamoja)
        df_all.to_excel(writer, sheet_name="ALL_DATA", index=False)
        style_worksheet(writer.sheets["ALL_DATA"], HEADER_COLORS["ALL_DATA"], df_all)
    
    # ── Muhtasari ─────────────────────────────────────────────────
    total = len(df_tcu) + len(df_nacte1) + len(df_nacte2)
    print("\n" + "=" * 60)
    print("  IMEKAMILIKA!")
    print(f"  Faili: {OUTPUT_FILE}")
    print(f"  TCU:      {len(df_tcu):>6,} rekodi")
    print(f"  NACTE_1:  {len(df_nacte1):>6,} rekodi")
    print(f"  NACTE_2:  {len(df_nacte2):>6,} rekodi")
    print(f"  JUMLA:    {total:>6,} rekodi")
    print("=" * 60)
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
