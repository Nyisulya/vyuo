"""
Management command: export_combined_excel
=========================================
Tumia: python manage.py export_combined_excel
       python manage.py export_combined_excel --source db
       python manage.py export_combined_excel --output /path/custom.xlsx

Mantiki:
  - Kama --source=excel: soma .xlsx files (TCU, NACTE1, NACTE2)
  - Kama --source=db (au excel files hazipo): soma kutoka database
  - Default VPS: itajaribu excel kwanza, kama hazipo → fallback kwa db
"""

import os
import sys
from pathlib import Path
from django.core.management.base import BaseCommand


# ── Rangi za header ──────────────────────────────────────────────────────────
HEADER_COLORS = {
    "TCU":           "1F4E79",
    "NACTE_1":       "375623",
    "NACTE_2":       "5C3317",
    "ALL_DATA":      "4B0082",
    "Database":      "2E4057",
}


def _style_ws(ws, hex_color, df):
    """Weka rangi, freeze, auto-filter, na auto-width kwenye worksheet."""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        fill   = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        font   = Font(bold=True, color="FFFFFF", size=11)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin   = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Header
        for cell in ws[1]:
            cell.fill, cell.font, cell.alignment, cell.border = fill, font, center, thin
        ws.row_dimensions[1].height = 28

        # Data rows
        for r_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            for cell in row:
                cell.alignment, cell.border = left, thin
                if r_idx % 2 == 0:
                    cell.fill = alt_fill
            ws.row_dimensions[r_idx].height = 18

        # Column widths
        for c_idx, col_name in enumerate(df.columns, 1):
            letter = get_column_letter(c_idx)
            if len(df) > 0:
                max_data = df[col_name].fillna("").astype(str).str.len().max()
                max_data = int(max_data) if max_data == max_data else 0
            else:
                max_data = 0
            ws.column_dimensions[letter].width = min(max(len(str(col_name)), max_data) + 4, 64)

        ws.freeze_panes = "A2"
        if ws.dimensions != "A1:A1":
            ws.auto_filter.ref = ws.dimensions

    except Exception:
        pass  # Styling ni optional, usisimame kwa hilo


class Command(BaseCommand):
    help = "Tengeneza Excel yenye data zote (TCU + NACTE) kutoka excel files au database"

    def add_arguments(self, parser):
        parser.add_argument(
            "--source",
            choices=["excel", "db", "auto"],
            default="auto",
            help=(
                "Chanzo cha data:\n"
                "  auto  → jaribu excel kwanza, fallback kwa db (DEFAULT)\n"
                "  excel → soma .xlsx files za TCU na NACTE\n"
                "  db    → soma kutoka database ya Django"
            ),
        )
        parser.add_argument(
            "--output",
            type=str,
            default=None,
            help="Njia kamili ya faili la kutoa (default: combined_data.xlsx)",
        )

    def handle(self, *args, **options):
        source = options["source"]
        output = options["output"]

        self.stdout.write("=" * 55)
        self.stdout.write("  KUTENGENEZA EXCEL - DATA ZOTE")
        self.stdout.write("=" * 55)

        try:
            import pandas as pd
        except ImportError:
            self.stderr.write(self.style.ERROR(
                "pandas haipo! Endesha: pip install pandas openpyxl"
            ))
            return

        # commands/ → management/ → universitysite/ → project_root/
        # parent×1     parent×2       parent×3          parent×4
        BASE_DIR = Path(__file__).resolve().parent.parent.parent.parent
        output_path = Path(output) if output else BASE_DIR / "combined_data.xlsx"
        self.stdout.write(f"  BASE_DIR: {BASE_DIR}")
        self.stdout.write(f"  Output:   {output_path}")

        # ── Chagua chanzo ────────────────────────────────────────────────────
        if source == "auto":
            tcu_path    = BASE_DIR / "tcu"   / "tcu_data.xlsx"
            nacte1_path = BASE_DIR / "nacte" / "nacte_data.xlsx"
            nacte2_path = BASE_DIR / "nacte" / "nacte_data2.xlsx"
            excel_exist = any([tcu_path.exists(), nacte1_path.exists(), nacte2_path.exists()])

            if excel_exist:
                self.stdout.write("  Chanzo: Excel files (.xlsx)")
                self._from_excel(pd, BASE_DIR, output_path)
            else:
                self.stdout.write("  Excel files hazipatikani → kutumia Database")
                self._from_db(pd, BASE_DIR, output_path)

        elif source == "excel":
            self._from_excel(pd, BASE_DIR, output_path)
        else:
            self._from_db(pd, BASE_DIR, output_path)

    # ════════════════════════════════════════════════════════════════════════
    # CHANZO 1: Excel files
    # ════════════════════════════════════════════════════════════════════════
    def _from_excel(self, pd, BASE_DIR, output_path):
        sys.path.insert(0, str(BASE_DIR))

        tcu_path    = BASE_DIR / "tcu"   / "tcu_data.xlsx"
        nacte1_path = BASE_DIR / "nacte" / "nacte_data.xlsx"
        nacte2_path = BASE_DIR / "nacte" / "nacte_data2.xlsx"

        self.stdout.write("\n[1/4] Kupakia TCU...")
        df_tcu = self._load_excel(pd, tcu_path, "TCU", "Degree")
        self.stdout.write(f"       Rekodi: {len(df_tcu):,}")

        self.stdout.write("[2/4] Kupakia NACTE_1...")
        df_n1 = self._load_excel(pd, nacte1_path, "NACTE_Book1", "Certificate/Diploma")
        self.stdout.write(f"       Rekodi: {len(df_n1):,}")

        self.stdout.write("[3/4] Kupakia NACTE_2...")
        df_n2 = self._load_excel(pd, nacte2_path, "NACTE_Book2", "Certificate/Diploma")
        self.stdout.write(f"       Rekodi: {len(df_n2):,}")

        # Kama zote ni empty, fallback kwa database
        if len(df_tcu) == 0 and len(df_n1) == 0 and len(df_n2) == 0:
            self.stdout.write(self.style.WARNING(
                "\n  [ONYO] Faili zote za Excel ni tupu au hazipo."
                "\n  Inabadilisha kwa Database..."
            ))
            self._from_db(pd, BASE_DIR, output_path)
            return

        self.stdout.write("[4/4] Kuandika Excel...")
        frames = [df for df in [df_tcu, df_n1, df_n2] if len(df) > 0]
        df_all = pd.concat(frames, ignore_index=True, sort=False).fillna("")

        # Panga columns: muhimu kwanza
        priority = ["Source", "Category", "University", "Programme", "Requirements", "Duration"]
        present_priority = [c for c in priority if c in df_all.columns]
        other = [c for c in df_all.columns if c not in present_priority]
        df_all = df_all[present_priority + other]

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            if len(df_tcu) > 0:
                df_tcu.to_excel(writer, sheet_name="TCU", index=False)
                _style_ws(writer.sheets["TCU"], HEADER_COLORS["TCU"], df_tcu)

            if len(df_n1) > 0:
                df_n1.to_excel(writer, sheet_name="NACTE_1", index=False)
                _style_ws(writer.sheets["NACTE_1"], HEADER_COLORS["NACTE_1"], df_n1)

            if len(df_n2) > 0:
                df_n2.to_excel(writer, sheet_name="NACTE_2", index=False)
                _style_ws(writer.sheets["NACTE_2"], HEADER_COLORS["NACTE_2"], df_n2)

            df_all.to_excel(writer, sheet_name="ALL_DATA", index=False)
            _style_ws(writer.sheets["ALL_DATA"], HEADER_COLORS["ALL_DATA"], df_all)

        total = len(df_tcu) + len(df_n1) + len(df_n2)
        self._done(output_path, total, tcu=len(df_tcu), n1=len(df_n1), n2=len(df_n2))

    def _load_excel(self, pd, path: Path, source_label: str, category: str) -> "pd.DataFrame":
        """Pakia .xlsx moja, ongeza Source na Category. Rudisha empty DF kama haipo."""
        if not path.exists():
            self.stdout.write(f"  [ONYO] Halipatikani: {path}")
            return pd.DataFrame()

        df = pd.read_excel(path)
        df["Source"]   = source_label
        df["Category"] = category

        # Normalize column names
        renames = {}
        for col in df.columns:
            cl = col.lower()
            if "university" in cl or "institution" in cl:
                renames[col] = "University"
            elif "programme" in cl or "program" in cl:
                renames[col] = "Programme"
            elif cl == "code":
                renames[col] = "Code"
            elif "requirement" in cl:
                renames[col] = "Requirements"
            elif "duration" in cl:
                renames[col] = "Duration"
            elif "capacity" in cl:
                renames[col] = "Capacity"
            elif cl == "fee":
                renames[col] = "Fee"
            elif "region" in cl:
                renames[col] = "Region"
            elif "ownership" in cl:
                renames[col] = "Ownership"
        df = df.rename(columns=renames)
        return df.fillna("")

    # ════════════════════════════════════════════════════════════════════════
    # CHANZO 2: Django Database
    # ════════════════════════════════════════════════════════════════════════
    def _from_db(self, pd, BASE_DIR, output_path):
        from universitysite.models import UniversityCourse

        self.stdout.write("\n[1/3] Kupakia data kutoka database...")

        qs = UniversityCourse.objects.select_related(
            "university", "course", "requirements", "university__region"
        ).order_by("university__name", "course__name")

        rows = []
        for uc in qs:
            rows.append({
                "Source":           "Database",
                "Category":         uc.level,
                "University":       uc.university.name,
                "Type":             uc.university.type or "",
                "Ownership":        uc.university.umiliki or "",
                "Region":           uc.university.region.name if uc.university.region_id else "",
                "Programme":        uc.course.name,
                "Level":            uc.level,
                "Duration":         uc.duration,
                "Requirements":     uc.requirements.description if uc.requirements else "",
                "Fee_TZS":          float(uc.fee) if uc.fee else "",
                "Application_Link": uc.application_link or "",
                "Is_Active":        uc.is_active,
            })

        df = pd.DataFrame(rows) if rows else pd.DataFrame()
        self.stdout.write(f"       Rekodi: {len(df):,}")

        if len(df) == 0:
            self.stdout.write(self.style.WARNING("  [ONYO] Database pia haina data!"))
            return

        self.stdout.write("[2/3] Kuandika Excel...")

        # Gawanya kwa level kwa sheets tofauti
        df_degree  = df[df["Level"] == "Degree"]
        df_diploma = df[df["Level"] == "Diploma"]
        df_cert    = df[df["Level"] == "Certificate"]

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Sheet kwa kila level
            if len(df_degree) > 0:
                df_degree.to_excel(writer, sheet_name="Degree", index=False)
                _style_ws(writer.sheets["Degree"], HEADER_COLORS["TCU"], df_degree)

            if len(df_diploma) > 0:
                df_diploma.to_excel(writer, sheet_name="Diploma", index=False)
                _style_ws(writer.sheets["Diploma"], HEADER_COLORS["NACTE_1"], df_diploma)

            if len(df_cert) > 0:
                df_cert.to_excel(writer, sheet_name="Certificate", index=False)
                _style_ws(writer.sheets["Certificate"], HEADER_COLORS["NACTE_2"], df_cert)

            # Sheet ya data yote
            df.to_excel(writer, sheet_name="ALL_DATA", index=False)
            _style_ws(writer.sheets["ALL_DATA"], HEADER_COLORS["ALL_DATA"], df)

        self._done(output_path, len(df), db=True,
                   degree=len(df_degree), diploma=len(df_diploma), cert=len(df_cert))

    # ════════════════════════════════════════════════════════════════════════
    def _done(self, path, total, tcu=0, n1=0, n2=0, db=False,
              degree=0, diploma=0, cert=0):
        self.stdout.write(self.style.SUCCESS("\n" + "=" * 55))
        self.stdout.write(self.style.SUCCESS("  IMEKAMILIKA!"))
        self.stdout.write(self.style.SUCCESS(f"  Faili:  {path}"))
        if db:
            self.stdout.write(self.style.SUCCESS(f"  Degree:      {degree:>6,} rekodi"))
            self.stdout.write(self.style.SUCCESS(f"  Diploma:     {diploma:>6,} rekodi"))
            self.stdout.write(self.style.SUCCESS(f"  Certificate: {cert:>6,} rekodi"))
        else:
            self.stdout.write(self.style.SUCCESS(f"  TCU:         {tcu:>6,} rekodi"))
            self.stdout.write(self.style.SUCCESS(f"  NACTE_1:     {n1:>6,} rekodi"))
            self.stdout.write(self.style.SUCCESS(f"  NACTE_2:     {n2:>6,} rekodi"))
        self.stdout.write(self.style.SUCCESS(f"  JUMLA:       {total:>6,} rekodi"))
        self.stdout.write(self.style.SUCCESS("=" * 55))
